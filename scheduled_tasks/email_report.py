import yagmail
from datetime import date
import os
import openpyxl
import json
import sys
from keys import GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD
from openpyxl import load_workbook

os.chdir(os.path.dirname(os.path.abspath(__file__)))

def email_report_daily():
    today_date = date.today().strftime("%m-%d-%y")
    report_path = f'../reports/dailyreports/{today_date}.xlsx'
    report = load_workbook(filename=report_path)
    worksheet = report.active
    if worksheet['A2'].value:
        yagmail.SMTP(GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD).send(
            to=['raunak@eabsentee.org', 'larry@eabsentee.org', 'robert@eabsentee.org'],
            subject=f'Daily Absentee Ballot Application Report - {today_date}',
            contents=f'Please find attached the daily report of absentee ' +
            f'ballot applications for {today_date}.',
            attachments=report_path
        )


def email_report(file_name, emails):
    today_date = date.today().strftime("%m-%d-%y")
    report = load_workbook(filename=file_name)
    worksheet = report.active
    if worksheet['B' + str(worksheet.max_row)].value.split()[0] == today_date:
        yagmail.SMTP(GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD).send(
            to=emails,
            cc=['raunak@eabsentee.org', 'larry@eabsentee.org'],
            subject=f'Daily Report - eAbsentee Applications',
            contents=f'New absentee ballot applications were submitted ' +
            'yesterday using your eAbsentee.org campaign link. Attached ' +
            'is a report on new and previously-submitted applications.',
            attachments=file_name
        )


def email_all_groups():
    with open('../static/groups.json') as file:
        groups_json = json.load(file)
        groups = groups_json.keys()
        for group in groups:
            if os.path.isfile('../reports/' + group + '.xlsx'):
                try:
                    email_report('../reports/' + group + '.xlsx',
                                 groups_json[group]['email'].split())
                except:
                    print('Oops, an error occurred with the group ' + group)


email_report_daily()
email_all_groups()
