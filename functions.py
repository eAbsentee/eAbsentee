import hashlib
import yagmail
import os
import sys
import openpyxl
import json
import datetime
import io
import googlemaps
from openpyxl import load_workbook
from flask import request, session
from datetime import date
from PyPDF2 import PdfFileWriter, PdfFileReader
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from subprocess import call
from gmplot import gmplot
from keys import GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD, API_KEY

# Change current working directory to directory 'functions.py' is in.
os.chdir(os.path.dirname(os.path.abspath(__file__)))

def application_process(request, group_code_form=None):
    data = parse_data(request, group_code_form=group_code_form)
    session_keys(data)
    write_pdf(data)

    try:
        build_report_data(data)
    except:
        yagmail.SMTP(GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD).send(
            to='raunak@eAbsentee.org',
            subject='Broken spreadsheet'
        )

    email_registrar(data)

    try:
        email_voter(data)
    except:
        pass

def parse_data(request, group_code_form):
    data = {}
    today_date = date.today().strftime('%m%d%y')

    group_code = ''
    if group_code_form is not None:
        group_code = group_code_form
    elif request.cookies.get('group'):
        group_code = request.cookies.get('group')

    emails_to_be_sent_to = []
    with open('static/localities_info.json') as file:
        localities = json.load(file)

        emails_to_be_sent_to = [localities[request.form['registered_county']]['email']] if 'Raunak Daga' not in str(str(request.form['name_first']) + ' ' + str(request.form['name_last'])) else []


        if request.form.get('email'):
            emails_to_be_sent_to.append(request.form.get('email'))

    # print(emails_to_be_sent_to)

    phonenumber = request.form.get('phonenumber').replace(
        '-', '').replace('(', '').replace(')', '').replace(' ', '').replace('+1', '').replace('-', '').replace('.', '').replace('+', '')

    with open('static/localities_info.json') as file:
        localities = json.load(file)
        data = {
            'first_name': request.form['name_first'],
            'middle_name': request.form['name_middle'],
            'last_name': request.form['name_last'],
            'suffix': request.form['name_suffix'],
            'full_name': request.form['name_first'] + ' ' + request.form['name_middle'] + ' ' + request.form['name_last'],
            'ssn': '  '.join(request.form['ssn']),
            'registered_to_vote': request.form['registered_county'] if request.form['name_first'] + ' ' + request.form['name_last'] != 'Raunak Daga' else 'Testing County',
            'email': request.form['email'],
            'address': request.form['address'],
            'apt': request.form['apt'],
            'city': request.form['city'],
            'zip_code': '   '.join(request.form['zip']),
            'state': 'VA',
            'full_address': request.form['address'] + ((' ' + request.form['apt']) if request.form['apt'] else ' ') + ', ' + request.form['city'] + ', ' + 'VA' +  ' ' + request.form['zip'],
            'full_delivery_address': request.form['different_address'] +      ((' ' + request.form['different_apt']) if request.form['different_apt'] else ' ') + ', ' + request.form['different_city'] + ', ' + request.form['different_state'] +  ' ' + request.form['different_zip'] + ' ' + request.form['different_country'],
            'delivery_address': request.form.get('different_address'),
            'delivery_city': request.form.get('different_city'),
            'delivery_apt': request.form.get('different_apt'),
            'delivery_zip': '   '.join(request.form.get( 'different_zip').replace('-', '')),
            'delivery_country': request.form['different_country'] if request.form['different_city'] else '',
            'delivery_state': request.form.get('different_state'),
            'signature': '/S/ ' + request.form[
                'signature'].replace('/S/', '', 1).strip(),
            'phonenumber': phonenumber,
            'assistant_check': 'X' if request.form.get(
                'assistance_check') == 'true' else '',
            'assistant_fullname': request.form.get('assistant_name'),
            'assistant_address': request.form.get('assistant_address'),
            'assistant_signature': request.form.get('assistant_name'),
            'assistant_apt': request.form.get('assistant_apt'),
            'assistant_city': request.form.get('assistant_city'),
            'assistant_state': request.form.get('assistant_state'),
            'assistant_zip': '   '.join(request.form.get(
                'assistant_zip').replace('-', '')),
            'deliver_residence': 'X' if request.form.get(
                'different_address_check') == 'false' else '',
            'deliver_mailing': 'X' if request.form.get(
                'different_address_check') == 'true' else '',
            'county_check': 'X' if 'County' in localities[
                request.form['registered_county']]['locality'] else '',
            'city_check': 'X' if 'City' in localities[
                request.form['registered_county']]['locality'] else '',
            'date_today_month': today_date[0:2],
            'date_today_day': today_date[2:4],
            'date_today_year': today_date[4:6],
            'application_ip': request.environ.get('HTTP_X_REAL_IP', request.remote_addr),
            'group_code': group_code,
            'registrar_address': localities[request.form['registered_county']]['email'],
            'emails_to_be_sent_to': emails_to_be_sent_to,
            'dem_prim_check': '',
            'rep_prim_check': '',
            'gen_spec_check': '',
            'date_election_year': '   '.join('20'),
            'date_election_day': '   '.join('03'),
            'date_election_month': '   '.join('11'),
            'full_election_date': '11 03 20'
        }

    return data

def session_keys(data):
    id = hashlib.md5(repr(data).encode('utf-8')).hexdigest()[: 10]
    today_date = date.today().strftime('%m-%d-%y')

    session['name'] = data['first_name'] + ' ' + data['middle_name'] + ' ' + data['last_name'] + (', ' + data['suffix'] if data['suffix'].strip() else '')
    session['application_id'] = id
    session['output_file'] = f'applications/{id}.pdf'
    session['registrar_locality'] = data['registered_to_vote']
    session['registrar_email'] = data['registrar_address']
    session['report_file'] = f'reports/{today_date}.xlsx'

def write_pdf(data):
    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=letter)
    can.drawString(180, 732, data['last_name'])
    can.drawString(410, 732, data['first_name'])
    can.drawString(190, 715, data['middle_name'])
    can.drawString(405, 715, data['suffix'])

    can.drawString(525, 698, (data['ssn'])) # SSN
    can.drawString(282, 683, 'X') # Gen/Spec
    # can.drawString(405, 683, 'X') # Dem Prim
    # can.drawString(503, 683, 'X') # Republican Primary
    can.drawString(212, 670, '11') # Month of Election
    can.drawString(261, 670, '03') # Day of Election
    can.drawString(310, 670, '20') # Year of Election
    can.drawString(428, 670, data['registered_to_vote']) # City/County of
    # can.drawString(409, 658, 'X') # Vote by Mail in All Elections Yes
    # can.drawString(442, 658, 'X') # Vote by Mail in All Elections No
    # can.drawString(126, 633, 'X') # Dem Primary Ballots
    # can.drawString(234, 633, 'X') # Rep Primary Ballots
    # can.drawString(331, 633, 'X') # No Primary Ballots

    can.drawString(165, 614, data['address'])
    can.drawString(535, 614, data['apt'])
    can.drawString(145, 596, data['city'])
    can.drawString(408, 596, ''.join(data['zip_code']))

    can.drawString(165, 565, data['delivery_address'])
    can.drawString(545, 565, data['delivery_apt'])
    can.drawString(145, 548, data['delivery_city'])
    can.drawString(322, 548, data['delivery_state'])
    can.drawString(418, 548, ''.join(data['delivery_zip']))
    can.drawString(558, 548, data['delivery_country'])

    # Phone Number
    # can.drawString(183, 525, '0     0     0')
    # can.drawString(268, 525, '0     0     0')
    # can.drawString(346, 525, '0     0     0     0')

    can.drawString(175, 509, data['email'])
    can.drawString(275, 115, data['signature'])
    can.drawString(485, 115, data['date_today_month'])
    can.drawString(525, 115, data['date_today_day'])
    can.drawString(565, 115, data['date_today_year'])

    can.save()
    packet.seek(0)
    new_pdf = PdfFileReader(packet)
    existing_pdf = PdfFileReader('static/pdf/new_blank_app.pdf', 'rb')
    output = PdfFileWriter()
    page = existing_pdf.getPage(0)
    page.mergePage(new_pdf.getPage(0))
    output.addPage(page)

    output.write(open(session['output_file'], 'wb'))



def email_registrar(data):
    """Email the form to the registrar of the applicant's locality. """
    yagmail.SMTP(GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD).send(
        to=([email for email in data['emails_to_be_sent_to']]),
        subject=('Absentee Ballot Request - Applicant-ID: ' +
        f'{session["application_id"]}'),
        contents='Please find attached an absentee ballot request ' +
        f'submitted on behalf of {session["name"]} - from eAbsentee.org',
        attachments=session['output_file']
    )

def email_voter(data):
    yagmail.SMTP(GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD).send(
        to=(data['email']),
        subject=('Absentee Ballot Request'),
        contents='Please find attached a copy of your absentee ballot request. This request will be digitally processed by your local registrar and you will receive an absentee ballot in the mail after some time. You can check on the status of your application by visiting the state elections website at https://www.elections.virginia.gov.',
        attachments=session['output_file']
    )

def build_report_data(data):
    personal_report_data = [
        data['full_name'],
        str(datetime.datetime.now().strftime('%m-%d-%y %H:%M:%S')),
        # data['reason_code'].replace(' ', ''),
        # data['supporting'],
        data['registered_to_vote'],
        data['email'],
        data['phonenumber'],
        data['full_address'],
        data['full_delivery_address'],
        data['application_ip'],
        session['application_id'],
        data['group_code'],
        data['registrar_address'],
        data['full_election_date']
    ]

    org_report_data = [
        data['full_name'],
        str(datetime.datetime.now().strftime('%m-%d-%y %H:%M:%S')),
        data['registered_to_vote'],
        data['email'],
        data['phonenumber'],
        data['full_address'],
        data['full_delivery_address'],
        session['application_id'],
        data['group_code'],
        data['full_election_date']
    ]


    append_to_report(personal_report_data, data['group_code'], org_report_data)

def append_to_report(personal_report_data, group_code, org_report_data):
    """Add a row to the Excel spreadsheet with data from the application.
    If the spreadsheet doesn't already exist, create it. """

    # APPENDING TO ALL TIME SPREADSHEET
    report_path = f'reports/all_time.xlsx'
    if not os.path.isfile(report_path):
        create_personal_report(report_path)
    report = load_workbook(filename=report_path)
    worksheet = report.active
    worksheet.append(personal_report_data)
    report.save(report_path)

    # APPENDING TO TODAYS SPREADSHEET
    today_date = date.today().strftime('%m-%d-%y')
    report_path = f'reports/dailyreports/{today_date}.xlsx'
    if not os.path.isfile(report_path):
        create_personal_report(report_path)
    report = load_workbook(filename=report_path)
    worksheet = report.active
    worksheet.append(personal_report_data)
    report.save(report_path)

    # APPENDING TO GROUP SPREADSHEET
    if group_code != '':
        report_path = f'reports/{group_code}.xlsx'
        if not os.path.isfile(report_path):
            create_org_report(report_path)
        report = load_workbook(filename=report_path)
        worksheet = report.active
        worksheet.append(org_report_data)
        report.save(report_path)


def create_personal_report(file_path):
    report = openpyxl.Workbook()
    sh = report.active
    sh['A1'] = 'Applicant Name'
    sh['B1'] = 'Time Submitted'
    sh['C1'] = 'Reason Code'
    sh['D1'] = 'Supporting Information'
    sh['E1'] = 'Locality'
    sh['F1'] = 'Email'
    sh['G1'] = 'Telephone Number'
    sh['H1'] = 'Address'
    sh['I1'] = 'Residence Address'
    sh['J1'] = 'IP Submitted From'
    sh['K1'] = 'Form ID'
    sh['L1'] = 'Group Code'
    sh['M1'] = 'Locality Email'
    sh['N1'] = 'Election Date'

    report.save(file_path)

def create_org_report(file_path):
    report = openpyxl.Workbook()
    sh = report.active
    sh['A1'] = 'Applicant Name'
    sh['B1'] = 'Time Submitted'
    sh['C1'] = 'Locality'
    sh['D1'] = 'Email'
    sh['E1'] = 'Telephone Number'
    sh['F1'] = 'Address'
    sh['G1'] = 'Residence Address'
    sh['H1'] = 'Form ID'
    sh['I1'] = 'Group Code'
    sh['J1'] = 'Election Date'

    report.save(file_path)


def add_to_groups(request):
    if request.form.get('api_key') != API_KEY:
        return

    if request.form.get('group_code'):
        with open('static/groups.json') as file:
            groups = json.load(file)

            new_group = None
            if request.form.get('county_codes'):
                new_group = {
                    request.form.get('group_code'): {
                        'email': request.form.get('group_email'),
                        'county_nums': request.form.get('county_codes').split()
                    }
                }
            else:
                new_group = {
                    request.form.get('group_code'): {
                        'email': request.form.get('group_email')
                    }
                }

            groups.update(new_group)
            with open('static/groups.json', 'w') as f:
                json.dump(groups, f, indent=4, sort_keys=True)


# Deprecated
def get_ids_and_counties(group_code):
    ids_and_names = {}
    with open('static/groups.json') as file:
        groups = json.load(file)
        group = ''
        # If a group has counties which it has selected to limit its form to
        if group_code in groups:
            group = groups[group_code]
            if 'county_nums' in group:
                with open('static/localities_info.json') as localities_file:
                    localities = json.load(localities_file)
                    for county_num_id in group['county_nums']:
                        ids_and_names[county_num_id] = localities[county_num_id]['locality']
                return ids_and_names
        # Otherwise, return all counties found in the group 'allcounties'
        with open('static/localities_info.json') as localities_file:
            localities = json.load(localities_file)
            for county_num_id in groups['allcounties']['county_nums']:
                ids_and_names[county_num_id] = localities[county_num_id]['locality']
        return ids_and_names



def email_report_alltime_api(request):
    report_path = f'reports/all_time.xlsx'
    report = load_workbook(filename=report_path)
    worksheet = report.active
    if request.form.get('api_key') == API_KEY:
        if worksheet['A2'].value:
            yagmail.SMTP(GMAIL_SENDER_ADDRESS, GMAIL_SENDER_PASSWORD).send(
                to=request.form.get('email_spreadsheet'),
                subject=f'Absentee Ballot Application Report',
                contents=f'Please find attached the report of absentee ' +
                f'ballot applications from all time.',
                attachments=f'reports/all_time.xlsx'
            )


def create_maps():
    MAPS_API_KEY = ''
    today_date = date.today().strftime('%m-%d-%y')
    report_path = f'reports/{today_date}.xlsx'
    report = load_workbook(filename=report_path)
    worksheet = report.active

    gmaps = googlemaps.Client(key=MAPS_API_KEY)

    gmap_plotter = gmplot.GoogleMapPlotter(40.3, -75, 6)
    gmap_plotter.apikey = MAPS_API_KEY

    for cell in worksheet['H']:
        str_value = str(cell.value)
        str_value = str_value.replace('   ', '')
        geocode_result = gmaps.geocode(str_value)[0]
        lat, lng = geocode_result['geometry']['location']['lat'], geocode_result['geometry']['location']['lng']
        gmap_plotter.marker(lat, lng)

    gmap_plotter.draw(f'maps/{today_date}.html')
